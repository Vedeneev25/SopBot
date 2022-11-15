
import telebot
from telebot import apihelper
import openpyxl
import cv2
import xlrd
import xlwt
from openpyxl import Workbook
from openpyxl import load_workbook


token = 

bot = telebot.TeleBot(token)

links = openpyxl.load_workbook(("Ссылки.xlsx"))
sheet1 = links['list1']
maxRow = sheet1.max_row
#Определяем все авиакомпании какие есть
print(maxRow)

def Airlines (sheet1, maxRow):
    airlines = []
    i = 1
    while i < maxRow:
        al = sheet1.cell(row = i, column = 1).value
        if al in airlines:
            i = i + 1
        else:
            airlines.append(al)
            i += 1
    #print (airlines)
    return airlines

#Определяем категории
def Categorii(Airlinename, maxRow, sheet1):
    #c = input("Выбери АК: ")
    category = []
    i = 1
    while i < maxRow:
        cat = sheet1.cell(row = i, column = 2).value
        if AirlineName == sheet1.cell(row = i, column = 1).value:
            if (cat in category):
                i = i + 1
            else:
                category.append(cat)
                i += 1
        else: 
            i +=1
        ##print (category)
    return category

#print(Categorii('Аэрофлот', maxRow, sheet1))
#print(Airlines(sheet1, maxRow))

AirlineName = ''
CategoryName = ''

@bot.message_handler(commands=['start'])
def start_message(message):
    bot.send_message(message.from_user.id, 'Привет! Я бот СОПП. Выбери авиакомпанию', reply_markup=keyboard2)
    bot.register_next_step_handler(message, category)


keyboard2 = telebot.types.ReplyKeyboardMarkup(True, True)
keyboard2.row('')
i = 0
while i < len(Airlines(sheet1, maxRow)):
        keyboard2.add(telebot.types.InlineKeyboardButton(text=Airlines(sheet1, maxRow)[i]))
        i = i + 1
 
def category(message):
    global AirlineName
    AirlineName = message.text
    keyboard1 = telebot.types.ReplyKeyboardMarkup(True, True)
    keyboard1.row('')
    i = 0
    while i < len(Categorii(AirlineName, maxRow, sheet1)):
        keyboard1.add(telebot.types.InlineKeyboardButton(text=Categorii(message, maxRow, sheet1)[i]))
        i = i + 1
    bot.send_message(message.from_user.id, 'Так, а теперь выбери категорию: ', reply_markup=keyboard1)
    bot.register_next_step_handler(message, send_text)



#@bot.message_handler(content_types=['text'])
def send_text(message):
    global CategoryName
    CategoryName = message.text
    i = 1
    while (i < maxRow):
        if (AirlineName == sheet1.cell(row = i, column = 1).value and CategoryName == sheet1.cell(row = i, column = 2).value):
            answer = sheet1.cell(row = i, column = 3).value
            break
        else:
            i += 1
    bot.send_message(message.from_user.id, "вот ссылка на ответ: " + answer)
        
        
    





bot.polling()

# def Airlines(message):
#     markup = telebot.types.InlineKeyboardMarkup()
#     markup.add(telebot.types.InlineKeyboardButton(text='Аэрофлот', callback_data=3))
#     bot.send_message(message.chat.id, text="выбери авиакомпанию", reply_markup=markup)

# @bot.message_handler(content_types=['text'])
# def send_text(message):
#     if message.text.lower() == 'привет':
#         bot.send_message(message.chat.id, 'Привет! ззз')
#         Airlines(message)

# @bot.callback_query_handler(func=lambda call: True)
# def query_handler(call,):
#     bot.answer_callback_query(callback_query_id=call.id, text='Приятного времени ожидания!')                                                                                       
#     answer = ''
#     if (call.data == sheet1.cell(row = i, column = 2).value and message.text == sheet1.cell(row = i, column = 1).value):
#         answer = 'link1'
#         bot.send_message(call.message.chat.id, answer, reply_markup=keyboard2)



# def AirlinesTG(message):
#     markup = telebot.types.InlineKeyboardMarkup()
#     i = 0
#     while i < len(Airlines(sheet1, maxRow)):
#         markup.add(telebot.types.InlineKeyboardButton(text=Airlines(sheet1, maxRow)[i]))
#         i = i + 1
#     bot.send_message(message.chat.id, 'Выбери авиакомпанию: ', reply_markup=keyboard2)
